from phovea_server.ns import Namespace, request, abort, Response
from phovea_server.util import jsonify
from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile
import logging

__author__ = 'Samuel Gratzl'
_log = logging.getLogger(__name__)
app = Namespace(__name__)


@app.route('/to_json', methods=['POST'])
def _xslx2json():
  file = request.files.get('file')
  if not file:
    abort(403, 'missing file')

  wb = load_workbook(file, read_only=True)

  def convert_sheet(ws):
    # ws.iter_rows()
    # ws.iter_cols()
    return dict(title=ws.title)

  data = dict(
    sheets=[convert_sheet(ws) for ws in wb.worksheets]
  )

  return jsonify(data)


@app.route('/from_json', methods=['POST'])
def _json2xslx():
  data = request.json
  wb = Workbook(write_only=True)

  for sheet in data.get('sheets', []):
    ws = wb.create_sheet(title=sheet['title'])

  with NamedTemporaryFile() as tmp:
    wb.save(tmp.name)
    tmp.seek(0)
    s = tmp.read()
    return Response(s, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



def create():
  """
   entry point of this plugin
  """
  return app
