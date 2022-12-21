import logging
from datetime import datetime
from tempfile import NamedTemporaryFile

import dateutil.parser
from flask import Flask, abort, jsonify, request
from flask.wrappers import Response
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font

_log = logging.getLogger(__name__)
app = Flask(__name__)


_types = dict(b="boolean", s="string")


def to_type(cell):
    if not cell:
        return "string"
    if cell.is_date:
        return "date"
    if cell.data_type in _types:
        return _types[cell.data_type]
    v = cell.value
    if isinstance(v, int) or isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    return "string"


def _convert_value(v):
    if isinstance(v, datetime):
        return v.isoformat()
    return v


@app.route("/to_json", methods=["POST"])
def _xlsx2json():
    file = request.files.get("file")
    if not file:
        abort(403, "missing file")

    wb = load_workbook(file, read_only=True, data_only=True)  # type: ignore

    def convert_row(row, cols):
        result = {}

        for r, c in zip(cols, row):
            result[c["name"]] = _convert_value(r.value)

        return result

    def convert_sheet(ws):

        ws_rows = ws.iter_rows()
        ws_cols = next(ws_rows, [])
        ws_first_row = next(ws_rows, [])

        cols = [dict(name=h.value, type=to_type(r)) for h, r in zip(ws_cols, ws_first_row)]

        rows = []
        rows.append(convert_row(cols, ws_first_row))
        for row in ws_rows:
            rows.append(str(convert_row(cols, row)))

        return dict(title=ws.title, columns=cols, rows=rows)

    data = dict(sheets=[convert_sheet(ws) for ws in wb.worksheets])

    return jsonify(data)


@app.route("/to_json_array", methods=["POST"])
def _xlsx2json_array():
    file = request.files.get("file")
    if not file:
        abort(403, "missing file")

    wb = load_workbook(file, read_only=True, data_only=True)  # type: ignore

    def convert_row(row):
        return [_convert_value(cell.value) for cell in row]

    if not wb.worksheets:
        return jsonify([])

    ws = wb.worksheets[0]

    rows = [convert_row(row) for row in ws.iter_rows()]
    return jsonify(rows)


@app.route("/from_json", methods=["POST"])
def _json2xlsx():
    data: dict = request.json  # type: ignore
    wb = Workbook(write_only=True)

    bold = Font(bold=True)

    def to_cell(v):
        # If the native value cannot be used as Excel value, used the stringified version instead.
        try:
            return WriteOnlyCell(ws, value=v)  # type: ignore
        except ValueError:
            return WriteOnlyCell(ws, value=str(v))  # type: ignore

    def to_header(v):
        c = to_cell(v)
        c.font = bold
        return c

    def to_value(v, coltype):
        if coltype == "date":
            if isinstance(v, int):
                v = datetime.fromtimestamp(v)
            elif isinstance(v, str) and len(v) > 0:
                v = dateutil.parser.parse(v)
        return to_cell(v)

    for sheet in data.get("sheets", []):
        ws = wb.create_sheet(title=sheet["title"])
        cols = sheet["columns"]
        ws.append(to_header(col["name"]) for col in cols)

        for row in sheet["rows"]:
            ws.append(to_value(row.get(col["name"], None), col["type"]) for col in cols)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        s = tmp.read()
        return Response(
            s,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


@app.route("/from_json_array", methods=["POST"])
def _json_array2xlsx():
    data: list = request.json  # type: ignore
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    for row in data:
        ws.append(row)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        s = tmp.read()
        return Response(
            s,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def create():
    """
    entry point of this plugin
    """
    return app
